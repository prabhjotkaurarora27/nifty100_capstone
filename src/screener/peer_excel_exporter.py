"""
src/screener/peer_excel_exporter.py
=====================================
Sprint 3 — Day 20
Export output/peer_comparison.xlsx: 11 sheets (one per peer group),
percentile colour-coded, benchmark row in gold, median summary row.

Run with:
    python -m src.screener.peer_excel_exporter
"""

from __future__ import annotations

import logging
import sqlite3
import sys
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
import config
from src.analytics.peer import PEER_METRICS

logger = logging.getLogger(__name__)

OUTPUT_PATH = config.OUTPUT_DIR / "peer_comparison.xlsx"

# ── 20 metric columns: (column_in_financial_ratios, display_header) ───────────
METRIC_COLS: List[tuple] = [
    ("return_on_equity_pct", "ROE (%)"),
    ("return_on_capital_employed_pct", "ROCE (%)"),
    ("net_profit_margin_pct", "NPM (%)"),
    ("operating_profit_margin_pct", "OPM (%)"),
    ("debt_to_equity", "D/E"),
    ("interest_coverage", "ICR"),
    ("icr_label", "ICR Label"),
    ("free_cash_flow_cr", "FCF (Cr)"),
    ("fcf_conversion_rate", "FCF Conv Rate"),
    ("cfo_quality_score", "CFO Quality"),
    ("revenue_cagr_3yr", "Rev CAGR 3yr (%)"),
    ("revenue_cagr_5yr", "Rev CAGR 5yr (%)"),
    ("pat_cagr_5yr", "PAT CAGR 5yr (%)"),
    ("eps_cagr_5yr", "EPS CAGR 5yr (%)"),
    ("asset_turnover", "Asset Turnover"),
    ("pe_ratio", "P/E"),
    ("pb_ratio", "P/B"),
    ("dividend_yield_pct", "Div Yield (%)"),
    ("market_cap_crore", "Mkt Cap (Cr)"),
    ("composite_quality_score", "Quality Score"),
]

# Percentile metric labels (from peer_percentiles table) matching PEER_METRICS
PEER_METRIC_LABELS = [label for _, label, _ in PEER_METRICS]

# ── Colour fills ──────────────────────────────────────────────────────────────
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")  # percentile >= 0.75
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")  # 0.25 <= pct < 0.75
FILL_RED = PatternFill("solid", fgColor="FFC7CE")  # percentile <= 0.25
FILL_GOLD = PatternFill("solid", fgColor="FFD700")  # benchmark row
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_SUBHEADER = PatternFill("solid", fgColor="2E75B6")
FILL_SUMMARY = PatternFill("solid", fgColor="E2EFDA")  # median summary row
FONT_WHITE_BOLD = Font(color="FFFFFF", bold=True, size=10)
FONT_HEADER_SM = Font(color="FFFFFF", bold=True, size=9)
FONT_BODY = Font(size=9)
FONT_BENCH_BOLD = Font(bold=True, size=9, color="1B1B1B")
FONT_SUMMARY = Font(italic=True, bold=True, size=9)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


def _percentile_fill(pr: Optional[float]) -> PatternFill:
    """Green >= 75th, Yellow 25–75th, Red <= 25th."""
    if pr is None or pd.isna(pr):
        return FILL_YELLOW
    if pr >= 0.75:
        return FILL_GREEN
    if pr <= 0.25:
        return FILL_RED
    return FILL_YELLOW


def _fmt(v) -> str:
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return "—"
    if isinstance(v, float):
        return f"{v:.2f}"
    return str(v)


# ─────────────────────────────────────────────────────────────────────────────
# Data loading
# ─────────────────────────────────────────────────────────────────────────────


def _load_all_data(db_path: Path) -> tuple:
    """Return (ratios_df, peer_groups_df, percentiles_df)."""
    metric_col_names = ", ".join(
        f"fr.{col}" for col, _ in METRIC_COLS if col not in ("icr_label",)
    )
    sql_ratios = f"""
    SELECT fr.company_id, c.company_name, fr.year,
           fr.icr_label,
           {metric_col_names}
    FROM financial_ratios fr
    JOIN companies c ON fr.company_id = c.id
    WHERE fr.year = (
        SELECT MAX(year) FROM financial_ratios fr2
        WHERE fr2.company_id = fr.company_id
          AND fr2.return_on_equity_pct IS NOT NULL
    )
    """
    sql_pg = "SELECT peer_group_name, company_id, is_benchmark FROM peer_groups"
    sql_pct = "SELECT company_id, peer_group_name, metric, percentile_rank FROM peer_percentiles"

    with sqlite3.connect(db_path) as conn:
        ratios_df = pd.read_sql_query(sql_ratios, conn)
        peer_groups_df = pd.read_sql_query(sql_pg, conn)
        try:
            percentiles_df = pd.read_sql_query(sql_pct, conn)
        except Exception:
            percentiles_df = pd.DataFrame(
                columns=["company_id", "peer_group_name", "metric", "percentile_rank"]
            )

    return ratios_df, peer_groups_df, percentiles_df


# ─────────────────────────────────────────────────────────────────────────────
# Sheet writer
# ─────────────────────────────────────────────────────────────────────────────


def _write_peer_sheet(
    ws,
    group_name: str,
    group_members: pd.DataFrame,  # peer_groups rows for this group
    ratios_df: pd.DataFrame,
    percentiles_df: pd.DataFrame,
) -> None:
    """Write one peer group sheet."""
    # Build pivoted percentiles: {company_id: {metric_label: percentile_rank}}
    pct_pivot: Dict[str, Dict[str, float]] = {}
    if not percentiles_df.empty:
        grp_pct = percentiles_df[percentiles_df["peer_group_name"] == group_name]
        for _, row in grp_pct.iterrows():
            pct_pivot.setdefault(row["company_id"], {})[row["metric"]] = row[
                "percentile_rank"
            ]

    # Columns = identity cols + metric cols + percentile rank cols per peer metric
    id_headers = ["Ticker", "Company", "Year"]
    metric_headers = [h for _, h in METRIC_COLS]
    rank_headers = [f"{lbl} Rank" for lbl in PEER_METRIC_LABELS]

    all_headers = id_headers + metric_headers + rank_headers
    n_cols = len(all_headers)

    # ── Row 1: Group title ────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 24
    title_cell = ws.cell(row=1, column=1, value=f"Peer Group: {group_name}")
    title_cell.font = Font(bold=True, size=13, color="1F4E79")
    title_cell.alignment = ALIGN_LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # ── Row 2: Section labels ─────────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    ws.cell(row=2, column=1, value="Identity").fill = FILL_HEADER
    ws.cell(row=2, column=1).font = FONT_WHITE_BOLD
    ws.cell(row=2, column=1).alignment = ALIGN_CENTER
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(id_headers))

    metric_start = len(id_headers) + 1
    ws.cell(row=2, column=metric_start, value="Key Metrics").fill = FILL_SUBHEADER
    ws.cell(row=2, column=metric_start).font = FONT_WHITE_BOLD
    ws.cell(row=2, column=metric_start).alignment = ALIGN_CENTER
    ws.merge_cells(
        start_row=2,
        start_column=metric_start,
        end_row=2,
        end_column=metric_start + len(metric_headers) - 1,
    )

    rank_start = metric_start + len(metric_headers)
    ws.cell(row=2, column=rank_start, value="Percentile Ranks").fill = FILL_SUBHEADER
    ws.cell(row=2, column=rank_start).font = FONT_WHITE_BOLD
    ws.cell(row=2, column=rank_start).alignment = ALIGN_CENTER
    ws.merge_cells(
        start_row=2,
        start_column=rank_start,
        end_row=2,
        end_column=rank_start + len(rank_headers) - 1,
    )

    # ── Row 3: Column headers ─────────────────────────────────────────────────
    ws.row_dimensions[3].height = 28
    for col_idx, header in enumerate(all_headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER_SM
        cell.alignment = ALIGN_CENTER

    # ── Data rows ─────────────────────────────────────────────────────────────
    benchmark_ids = set(
        group_members.loc[group_members["is_benchmark"] == 1, "company_id"]
    )
    company_ids = group_members["company_id"].tolist()

    # Collect numeric data for median
    numeric_data: Dict[str, List[float]] = {
        h: [] for h in metric_headers + rank_headers
    }

    data_start_row = 4
    for row_offset, cid in enumerate(company_ids):
        row_num = data_start_row + row_offset
        ws.row_dimensions[row_num].height = 16

        ratio_row = ratios_df[ratios_df["company_id"] == cid]
        is_bench = cid in benchmark_ids

        # Identity columns
        ticker_cell = ws.cell(row=row_num, column=1, value=cid)
        name_cell = ws.cell(
            row=row_num,
            column=2,
            value=ratio_row["company_name"].iloc[0] if not ratio_row.empty else cid,
        )
        year_cell = ws.cell(
            row=row_num,
            column=3,
            value=int(ratio_row["year"].iloc[0]) if not ratio_row.empty else None,
        )

        for c in (ticker_cell, name_cell, year_cell):
            c.font = FONT_BENCH_BOLD if is_bench else FONT_BODY
            c.alignment = ALIGN_LEFT if c.column == 2 else ALIGN_CENTER
            if is_bench:
                c.fill = FILL_GOLD

        # Metric columns
        for m_idx, (col_key, m_header) in enumerate(METRIC_COLS):
            col_num = metric_start + m_idx
            raw = (
                ratio_row[col_key].iloc[0]
                if not ratio_row.empty and col_key in ratio_row.columns
                else None
            )
            cell = ws.cell(row=row_num, column=col_num, value=_fmt(raw))
            cell.font = FONT_BENCH_BOLD if is_bench else FONT_BODY
            cell.alignment = ALIGN_CENTER
            if is_bench:
                cell.fill = FILL_GOLD
            # Collect numeric
            try:
                numeric_data[m_header].append(float(raw))
            except (TypeError, ValueError):
                pass

        # Percentile rank columns
        cid_pct = pct_pivot.get(cid, {})
        for r_idx, lbl in enumerate(PEER_METRIC_LABELS):
            col_num = rank_start + r_idx
            pr = cid_pct.get(lbl)
            display = f"{pr:.3f}" if pr is not None else "—"
            cell = ws.cell(row=row_num, column=col_num, value=display)
            cell.alignment = ALIGN_CENTER
            cell.font = FONT_BENCH_BOLD if is_bench else FONT_BODY
            if is_bench:
                cell.fill = FILL_GOLD
            else:
                cell.fill = _percentile_fill(pr)
            # Collect numeric
            if pr is not None:
                try:
                    numeric_data[f"{lbl} Rank"].append(float(pr))
                except (TypeError, ValueError):
                    pass

    # ── Summary (median) row ──────────────────────────────────────────────────
    summary_row = data_start_row + len(company_ids)
    ws.row_dimensions[summary_row].height = 18
    ws.cell(row=summary_row, column=1, value="Peer Median").fill = FILL_SUMMARY
    ws.cell(row=summary_row, column=1).font = FONT_SUMMARY
    ws.cell(row=summary_row, column=1).alignment = ALIGN_CENTER
    ws.cell(row=summary_row, column=2, value="").fill = FILL_SUMMARY
    ws.cell(row=summary_row, column=3, value="").fill = FILL_SUMMARY

    for m_idx, (_, m_header) in enumerate(METRIC_COLS):
        col_num = metric_start + m_idx
        vals = numeric_data.get(m_header, [])
        med = float(pd.Series(vals).median()) if vals else None
        cell = ws.cell(row=summary_row, column=col_num, value=_fmt(med))
        cell.fill = FILL_SUMMARY
        cell.font = FONT_SUMMARY
        cell.alignment = ALIGN_CENTER

    for r_idx, lbl in enumerate(PEER_METRIC_LABELS):
        col_num = rank_start + r_idx
        vals = numeric_data.get(f"{lbl} Rank", [])
        med = float(pd.Series(vals).median()) if vals else None
        cell = ws.cell(row=summary_row, column=col_num, value=_fmt(med))
        cell.fill = FILL_SUMMARY
        cell.font = FONT_SUMMARY
        cell.alignment = ALIGN_CENTER

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 12
    ws.column_dimensions[get_column_letter(2)].width = 28
    ws.column_dimensions[get_column_letter(3)].width = 7
    for i in range(len(metric_headers)):
        ws.column_dimensions[get_column_letter(metric_start + i)].width = 14
    for i in range(len(rank_headers)):
        ws.column_dimensions[get_column_letter(rank_start + i)].width = 12


# ─────────────────────────────────────────────────────────────────────────────
# Public export function
# ─────────────────────────────────────────────────────────────────────────────


def export_peer_comparison_xlsx(
    db_path: Path = config.DB_PATH,
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """
    Write output/peer_comparison.xlsx with exactly 11 sheets (one per peer group).

    Returns
    -------
    Path : the written file path
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    ratios_df, peer_groups_df, percentiles_df = _load_all_data(db_path)

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    group_names = sorted(peer_groups_df["peer_group_name"].unique())
    for group_name in group_names:
        members = peer_groups_df[peer_groups_df["peer_group_name"] == group_name]
        ws = wb.create_sheet(title=group_name[:31])
        _write_peer_sheet(ws, group_name, members, ratios_df, percentiles_df)
        logger.info("Sheet '%s' written (%d companies)", group_name, len(members))

    wb.save(output_path)
    logger.info(
        "✓ peer_comparison.xlsx saved → %s  (%d sheets)", output_path, len(group_names)
    )
    return output_path


# ─────────────────────────────────────────────────────────────────────────────
# CLI entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S",
    )
    path = export_peer_comparison_xlsx()
    print(f"\n✓ Exported: {path}")
    import openpyxl

    wb = openpyxl.load_workbook(path)
    print(f"  Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")
