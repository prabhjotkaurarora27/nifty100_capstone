"""
src/screener/excel_exporter.py
================================
Sprint 3 — Day 17
Export screener_output.xlsx: one sheet per preset, 20 KPI columns,
sorted by composite_quality_score descending, green/red colour-coded.

Run with:
    python -m src.screener.excel_exporter
"""

from __future__ import annotations

import logging
import sys
from pathlib import Path
from typing import Dict, List, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
import config

logger = logging.getLogger(__name__)

OUTPUT_PATH = config.OUTPUT_DIR / "screener_output.xlsx"

# ── 20 KPI columns to show (column_key, display_header) ──────────────────────
KPI_DISPLAY_COLS: List[tuple] = [
    ("company_id", "Ticker"),
    ("company_name", "Company"),
    ("broad_sector", "Sector"),
    ("year", "Year"),
    ("composite_quality_score", "Quality Score"),
    ("sector_relative_score", "Sector Score"),
    ("return_on_equity_pct", "ROE (%)"),
    ("return_on_capital_employed_pct", "ROCE (%)"),
    ("net_profit_margin_pct", "NPM (%)"),
    ("operating_profit_margin_pct", "OPM (%)"),
    ("debt_to_equity", "D/E"),
    ("interest_coverage", "ICR"),
    ("free_cash_flow_cr", "FCF (Cr)"),
    ("fcf_conversion_rate", "FCF Conv Rate"),
    ("cfo_quality_score", "CFO Quality"),
    ("revenue_cagr_5yr", "Rev CAGR 5yr (%)"),
    ("pat_cagr_5yr", "PAT CAGR 5yr (%)"),
    ("eps_cagr_5yr", "EPS CAGR 5yr (%)"),
    ("pe_ratio", "P/E"),
    ("pb_ratio", "P/B"),
    ("dividend_yield_pct", "Div Yield (%)"),
    ("market_cap_crore", "Mkt Cap (Cr)"),
    ("asset_turnover", "Asset Turnover"),
]

# ── Thresholds for green (pass) / red (fail) per column ──────────────────────
# Format: (column_key, direction, threshold)
#   direction "min" → value >= threshold is green
#   direction "max" → value <= threshold is green
THRESHOLD_RULES: List[tuple] = [
    ("return_on_equity_pct", "min", 15.0),
    ("return_on_capital_employed_pct", "min", 12.0),
    ("net_profit_margin_pct", "min", 8.0),
    ("operating_profit_margin_pct", "min", 10.0),
    ("debt_to_equity", "max", 1.0),
    ("interest_coverage", "min", 2.0),
    ("free_cash_flow_cr", "min", 0.0),
    ("fcf_conversion_rate", "min", 0.5),
    ("revenue_cagr_5yr", "min", 10.0),
    ("pat_cagr_5yr", "min", 10.0),
    ("eps_cagr_5yr", "min", 8.0),
    ("pe_ratio", "max", 35.0),
    ("pb_ratio", "max", 5.0),
    ("dividend_yield_pct", "min", 0.5),
    ("asset_turnover", "min", 0.5),
    ("composite_quality_score", "min", 40.0),
]
_THRESHOLD_MAP = {
    col: (direction, thresh) for col, direction, thresh in THRESHOLD_RULES
}

# ── Excel style constants ─────────────────────────────────────────────────────
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FONT_HEADER = Font(color="FFFFFF", bold=True, size=10)
FONT_BODY = Font(size=9)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")


def _cell_fill(col_key: str, value) -> Optional[PatternFill]:
    """Return green fill if value passes threshold, red if fails, None otherwise."""
    if col_key not in _THRESHOLD_MAP:
        return None
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    direction, threshold = _THRESHOLD_MAP[col_key]
    try:
        v = float(value)
    except (TypeError, ValueError):
        return None
    if direction == "min":
        return FILL_GREEN if v >= threshold else FILL_RED
    else:  # max
        return FILL_GREEN if v <= threshold else FILL_RED


def _fmt_value(value) -> str:
    """Format a cell value for display."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "—"
    if isinstance(value, float):
        return f"{value:.2f}"
    return str(value)


def _write_preset_sheet(ws, preset_name: str, df: pd.DataFrame) -> None:
    """Write one preset's data to an openpyxl worksheet."""
    # Filter to available columns only
    avail_cols = [(k, h) for k, h in KPI_DISPLAY_COLS if k in df.columns]

    # ── Header row ────────────────────────────────────────────────────────────
    ws.row_dimensions[1].height = 30
    for col_idx, (col_key, header) in enumerate(avail_cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = ALIGN_CENTER

    # ── Preset title in a merged header above ─────────────────────────────────
    ws.insert_rows(1)
    title_cell = ws.cell(row=1, column=1, value=f"Nifty 100 Screener — {preset_name}")
    title_cell.font = Font(bold=True, size=12, color="1F4E79")
    title_cell.alignment = ALIGN_LEFT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(avail_cols))
    ws.row_dimensions[1].height = 22

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_idx, (_, row) in enumerate(df.iterrows(), start=3):
        ws.row_dimensions[row_idx].height = 16
        for col_idx, (col_key, _) in enumerate(avail_cols, start=1):
            raw = row.get(col_key) if col_key in df.columns else None
            display = _fmt_value(raw)
            cell = ws.cell(row=row_idx, column=col_idx, value=display)
            cell.font = FONT_BODY
            cell.alignment = ALIGN_CENTER

            fill = _cell_fill(col_key, raw)
            if fill:
                cell.fill = fill

    # ── Column widths ─────────────────────────────────────────────────────────
    col_widths = {
        "Ticker": 10,
        "Company": 28,
        "Sector": 20,
        "Year": 6,
        "CFO Quality": 14,
    }
    for col_idx, (col_key, header) in enumerate(avail_cols, start=1):
        width = col_widths.get(header, 14)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # ── Summary count ─────────────────────────────────────────────────────────
    summary_row = len(df) + 4
    ws.cell(row=summary_row, column=1, value=f"Total: {len(df)} companies").font = Font(
        bold=True, italic=True, size=9
    )


def export_screener_xlsx(
    preset_results: Dict[str, pd.DataFrame],
    output_path: Path = OUTPUT_PATH,
) -> Path:
    """
    Write output/screener_output.xlsx with one sheet per preset.

    Parameters
    ----------
    preset_results : dict of {preset_name: filtered DataFrame}
    output_path    : destination path

    Returns
    -------
    Path : the written file path
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for preset_name, df in preset_results.items():
        # Sheet name: max 31 chars, no invalid chars
        sheet_name = preset_name[:31]
        ws = wb.create_sheet(title=sheet_name)
        _write_preset_sheet(ws, preset_name, df)
        logger.info("Sheet '%s' written — %d companies", sheet_name, len(df))

    wb.save(output_path)
    logger.info("✓ screener_output.xlsx saved → %s", output_path)
    return output_path


# ── CLI entry point ────────────────────────────────────────────────────────────
if __name__ == "__main__":
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s  %(levelname)-8s  %(message)s",
        datefmt="%H:%M:%S",
    )
    from src.screener.engine import ScreenerEngine

    engine = ScreenerEngine()
    engine.load_data()
    results = engine.run_all_presets()
    path = export_screener_xlsx(results)
    print(f"✓ Exported: {path}")
