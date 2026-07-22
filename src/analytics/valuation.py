import sqlite3
from pathlib import Path
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
import pandas as pd

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DB_PATH = PROJECT_ROOT / "db" / "nifty100.db"
OUTPUT_DIR = PROJECT_ROOT / "output"


def run_valuation_analysis():
    """Calculates valuation metrics, sector medians, FCF yield, 5yr median PE,

    PE vs sector median %, overvaluation flags, and exports Excel & CSV reports.
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    conn = sqlite3.connect(str(DB_PATH))

    # Fetch 5-year historical ratios for 5yr median P/E computation
    hist_query = """
        SELECT fr.company_id, fr.year, fr.pe_ratio
        FROM financial_ratios fr
        WHERE fr.pe_ratio IS NOT NULL AND fr.pe_ratio > 0
    """
    hist_df = pd.read_sql_query(hist_query, conn)
    pe_5yr_median = (
        hist_df.groupby("company_id")["pe_ratio"].median().to_dict()
    )

    # Fetch latest year (2024) valuation data
    query = """
        SELECT fr.company_id, c.company_name, s.broad_sector as sector,
               fr.market_cap_crore, fr.pe_ratio, fr.pb_ratio, fr.ev_ebitda,
               fr.free_cash_flow_cr
        FROM financial_ratios fr
        JOIN companies c ON fr.company_id = c.id
        LEFT JOIN sectors s ON c.id = s.company_id
        WHERE fr.year = (SELECT MAX(year) FROM financial_ratios)
        ORDER BY c.company_name ASC
    """
    df = pd.read_sql_query(query, conn)
    conn.close()

    # 1. Compute FCF Yield (%)
    df["FCF_yield_pct"] = (
        df["free_cash_flow_cr"] / df["market_cap_crore"]
    ) * 100

    # 2. Sector Median P/E per broad_sector
    sector_pe_medians = df.groupby("sector")["pe_ratio"].median().to_dict()
    df["sector_median_PE"] = df["sector"].map(sector_pe_medians)

    # 3. 5-year Median P/E per company
    df["5yr_median_PE"] = df["company_id"].map(pe_5yr_median)

    # 4. P/E vs Sector Median (%)
    df["PE_vs_sector_median_pct"] = (
        (df["pe_ratio"] - df["sector_median_PE"]) / df["sector_median_PE"]
    ) * 100

    # 5. Overvaluation Flags
    def get_flag(row):
        pe = row["pe_ratio"]
        sec_med = row["sector_median_PE"]
        if pd.isnull(pe) or pd.isnull(sec_med) or sec_med <= 0:
            return "Fair"
        if pe > sec_med * 1.5:
            return "Caution"
        elif pe < sec_med * 0.7:
            return "Discount"
        else:
            return "Fair"

    df["flag"] = df.apply(get_flag, axis=1)

    # Prepare output DataFrames
    summary_cols = [
        "company_id",
        "company_name",
        "sector",
        "pe_ratio",
        "pb_ratio",
        "ev_ebitda",
        "FCF_yield_pct",
        "5yr_median_PE",
        "PE_vs_sector_median_pct",
        "flag",
    ]

    summary_df = df[summary_cols].rename(
        columns={
            "pe_ratio": "P/E",
            "pb_ratio": "P/B",
            "ev_ebitda": "EV/EBITDA",
        }
    )

    # Export output/valuation_flags.csv (only Caution and Discount)
    flags_df = summary_df[summary_df["flag"].isin(["Caution", "Discount"])].copy()
    flags_path = OUTPUT_DIR / "valuation_flags.csv"
    flags_df.to_csv(flags_path, index=False)
    print(
        f"✅ Exported {len(flags_df)} flagged companies to {flags_path.name}"
    )

    # Export output/valuation_summary.xlsx
    excel_path = OUTPUT_DIR / "valuation_summary.xlsx"
    export_styled_excel(summary_df, excel_path)
    print(f"✅ Exported {len(summary_df)} rows to {excel_path.name}")

    return summary_df, flags_df


def export_styled_excel(df: pd.DataFrame, file_path: Path):
    """Exports valuation summary DataFrame to styled Excel workbook with conditional formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Valuation Summary"
    ws.views.sheetView[0].showGridLines = True

    header_fill = PatternFill(
        start_color="1F497D", end_color="1F497D", fill_type="solid"
    )
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    # Fills for Flags
    caution_fill = PatternFill(
        start_color="F8D7DA", end_color="F8D7DA", fill_type="solid"
    )
    caution_font = Font(name="Calibri", size=10, color="721C24", bold=True)
    discount_fill = PatternFill(
        start_color="D4EDDA", end_color="D4EDDA", fill_type="solid"
    )
    discount_font = Font(name="Calibri", size=10, color="155724", bold=True)
    fair_fill = PatternFill(
        start_color="E2E3E5", end_color="E2E3E5", fill_type="solid"
    )
    fair_font = Font(name="Calibri", size=10, color="383D41")

    # Header Row
    headers = list(df.columns)
    ws.append(headers)
    for col_num, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align

    # Data Rows
    for row in df.itertuples(index=False):
        ws.append(list(row))

    for r_idx, row_data in enumerate(
        df.itertuples(index=False), start=2
    ):
        for c_idx, val in enumerate(row_data, start=1):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.font = data_font
            cell.border = thin_border

            # Column specific formatting
            col_name = headers[c_idx - 1]
            if col_name in [
                "P/E",
                "P/B",
                "EV/EBITDA",
                "FCF_yield_pct",
                "5yr_median_PE",
                "PE_vs_sector_median_pct",
            ]:
                cell.alignment = right_align
                if pd.notnull(val):
                    cell.number_format = "#,##0.00"
                else:
                    cell.value = "N/A"
                    cell.alignment = center_align
            elif col_name in ["company_id"]:
                cell.alignment = center_align
            elif col_name == "flag":
                cell.alignment = center_align
                if val == "Caution":
                    cell.fill = caution_fill
                    cell.font = caution_font
                elif val == "Discount":
                    cell.fill = discount_fill
                    cell.font = discount_font
                else:
                    cell.fill = fair_fill
                    cell.font = fair_font
            else:
                cell.alignment = left_align

    # Auto-adjust column widths
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

    wb.save(file_path)


if __name__ == "__main__":
    run_valuation_analysis()
