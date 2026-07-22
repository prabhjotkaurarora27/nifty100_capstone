"""
src/etl/file_inspector.py
-------------------------
Scans data/raw/ and prints a summary table of every file:
  filename | size | sheets (xlsx) | rows | columns | sample column names

Useful before the first load to verify filenames and column headers
so FILE_MAP in loader.py can be adjusted accordingly.

Usage
-----
    python src/etl/file_inspector.py
    make inspect
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

# ── project root ──────────────────────────────────────────────────────────────
BASE_DIR: Path = Path(__file__).resolve().parents[2]
sys.path.insert(0, str(BASE_DIR))

import config as cfg

DATA_RAW_DIR: Path = cfg.DATA_RAW_DIR


# ── helpers ────────────────────────────────────────────────────────────────────


def _fmt_size(n_bytes: int) -> str:
    for unit in ("B", "KB", "MB", "GB"):
        if n_bytes < 1024:
            return f"{n_bytes:.1f} {unit}"
        n_bytes //= 1024
    return f"{n_bytes:.1f} TB"


def _inspect_xlsx(path: Path) -> dict:
    """Peek at an xlsx without loading the whole file."""
    try:
        import openpyxl

        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        sheets = wb.sheetnames
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        header = next(rows_iter, ())
        # count remaining rows quickly
        row_count = sum(1 for _ in rows_iter)
        wb.close()
        col_names = [str(h).strip() if h is not None else "" for h in header]
        return {
            "sheets": sheets,
            "rows": row_count,  # data rows (excl. header)
            "cols": len(col_names),
            "col_names": col_names,
            "error": None,
        }
    except Exception as exc:
        return {"sheets": [], "rows": 0, "cols": 0, "col_names": [], "error": str(exc)}


def _inspect_csv(path: Path) -> dict:
    """Count rows and read header from a CSV."""
    try:
        with path.open(encoding="utf-8", errors="replace") as f:
            reader = csv.reader(f)
            header = next(reader, [])
            row_count = sum(1 for _ in reader)
        return {
            "sheets": ["(csv)"],
            "rows": row_count,
            "cols": len(header),
            "col_names": header,
            "error": None,
        }
    except Exception as exc:
        return {"sheets": [], "rows": 0, "cols": 0, "col_names": [], "error": str(exc)}


def inspect_directory(directory: Path = DATA_RAW_DIR) -> list[dict]:
    """Scan directory and return a list of file-info dicts."""
    if not directory.exists():
        print(f"  ⚠️  Directory not found: {directory}")
        return []

    files = sorted(
        [
            p
            for p in directory.iterdir()
            if p.is_file() and p.suffix.lower() in (".xlsx", ".xls", ".csv")
        ],
        key=lambda p: p.name.lower(),
    )

    if not files:
        print(f"  ⚠️  No Excel/CSV files found in {directory}")
        return []

    results = []
    for path in files:
        size = path.stat().st_size
        suffix = path.suffix.lower()
        if suffix in (".xlsx", ".xls"):
            info = _inspect_xlsx(path)
        else:
            info = _inspect_csv(path)

        info["filename"] = path.name
        info["size"] = size
        results.append(info)

    return results


def _print_report(results: list[dict]) -> None:
    """Pretty-print the inspection table."""
    if not results:
        return

    # ── header ────────────────────────────────────────────────────────────────
    name_w = max(len(r["filename"]) for r in results) + 2
    sheet_w = max(len(", ".join(r["sheets"])) for r in results) + 2
    sheet_w = max(sheet_w, 12)
    col_sample_w = 60

    sep = "=" * (name_w + sheet_w + 7 + 7 + col_sample_w + 10)
    print()
    print(sep)
    print(
        f"  {'Filename':<{name_w}}  {'Sheet(s)':<{sheet_w}}  {'Rows':>6}  {'Cols':>5}  Column names (first 5)"
    )
    print(sep)

    for r in results:
        name = r["filename"]
        sheets = ", ".join(r["sheets"]) if r["sheets"] else "—"
        rows = r["rows"]
        cols = r["cols"]
        sample = ", ".join(r["col_names"][:5])
        size_s = _fmt_size(r["size"])
        err = r.get("error")

        if err:
            print(f"  {name:<{name_w}}  ❌ ERROR: {err}")
        else:
            print(
                f"  {name:<{name_w}}  {sheets:<{sheet_w}}  {rows:>6}  {cols:>5}  {sample}"
            )
            if len(r["col_names"]) > 5:
                extra = ", ".join(r["col_names"][5:10])
                ellipsis = "…" if len(r["col_names"]) > 10 else ""
                print(
                    f"  {'':>{name_w}}  {'':>{sheet_w}}  {'':>6}  {'':>5}  ({extra}{ellipsis})"
                )
        print(f"  {'':>{name_w}}  size: {size_s}")
        print()

    print(sep)
    print(f"  Total files: {len(results)}")
    print(sep)
    print()


def _print_loader_map_hints(results: list[dict]) -> None:
    """Print a quick-reference mapping of filename → detected columns."""
    from src.etl.loader import SOURCE_FILES

    expected = {e["filename"] for e in SOURCE_FILES}
    found = {r["filename"] for r in results}

    missing = expected - found
    extra = found - expected

    if missing:
        print("⚠️  Files expected by loader.py but NOT found in data/raw/:")
        for f in sorted(missing):
            print(f"     ✗  {f}")
        print()

    if extra:
        print("ℹ️  Extra files in data/raw/ not in loader SOURCE_FILES:")
        for f in sorted(extra):
            print(f"     +  {f}")
        print()

    if not missing and not extra:
        print("✅  All expected source files are present in data/raw/\n")


# ── entry point ────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    print(f"\n  Inspecting: {DATA_RAW_DIR}\n")
    results = inspect_directory(DATA_RAW_DIR)
    _print_report(results)
    if results:
        _print_loader_map_hints(results)
