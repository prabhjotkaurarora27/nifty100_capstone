"""
tests/screener/test_screener.py
================================
Sprint 3 — Day 21
14 DQ unit tests for the screener engine, composite score, and peer analytics.
All 14 tests must pass with 0 failures.

Run with:
    python -m pytest tests/screener/ -v
"""

from __future__ import annotations

import sqlite3
from pathlib import Path

import pandas as pd
import pytest
import yaml

# ─────────────────────────────────────────────────────────────────────────────
# Fixtures
# ─────────────────────────────────────────────────────────────────────────────

@pytest.fixture(scope="module")
def engine():
    """Shared ScreenerEngine instance (loads data once for the module)."""
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    from src.screener.engine import ScreenerEngine
    e = ScreenerEngine()
    e.load_data()
    return e


@pytest.fixture(scope="module")
def db_path() -> Path:
    import sys
    sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
    import config
    return config.DB_PATH


@pytest.fixture(scope="module")
def all_presets(engine):
    """Run all 6 presets once and cache results."""
    from src.screener.engine import ScreenerEngine
    return {name: engine.run_preset(name) for name in ScreenerEngine.PRESET_NAMES}


# ─────────────────────────────────────────────────────────────────────────────
# Test 1 — Config YAML loads without error
# ─────────────────────────────────────────────────────────────────────────────
class TestConfig:
    """Test 1"""

    def test_yaml_loads(self):
        """screener_config.yaml must parse without error and contain 'filters'."""
        config_path = (
            Path(__file__).resolve().parents[2] / "config" / "screener_config.yaml"
        )
        assert config_path.exists(), f"Config not found: {config_path}"
        with open(config_path, encoding="utf-8") as fh:
            cfg = yaml.safe_load(fh)
        assert isinstance(cfg, dict), "YAML must be a mapping at top level"
        assert "filters" in cfg, "YAML must contain 'filters' key"
        filters = cfg["filters"]
        # Check at least 10 filter keys present
        assert len(filters) >= 10, f"Expected >= 10 filters, got {len(filters)}"


# ─────────────────────────────────────────────────────────────────────────────
# Test 2 — Engine loads DB data correctly
# ─────────────────────────────────────────────────────────────────────────────
class TestEngineDataLoad:
    """Test 2"""

    def test_load_returns_dataframe_with_expected_columns(self, engine):
        """Engine.load_data() must return a DataFrame with key columns."""
        df = engine.data
        assert isinstance(df, pd.DataFrame), "data must be a DataFrame"
        assert len(df) > 0, "DataFrame must not be empty"
        required_cols = [
            "company_id", "company_name", "broad_sector",
            "return_on_equity_pct", "debt_to_equity",
            "composite_quality_score", "sector_relative_score",
        ]
        for col in required_cols:
            assert col in df.columns, f"Missing column: {col}"


# ─────────────────────────────────────────────────────────────────────────────
# Test 3 — Quality Compounder returns non-empty DataFrame
# ─────────────────────────────────────────────────────────────────────────────
class TestQualityCompounderNotEmpty:
    """Test 3"""

    def test_quality_compounder_not_empty(self, engine):
        """Quality Compounder preset must return at least 1 row."""
        result = engine.preset_quality_compounder()
        assert len(result) > 0, "Quality Compounder returned 0 results"


# ─────────────────────────────────────────────────────────────────────────────
# Test 4 — Quality Compounder: all results ROE > 15%
# ─────────────────────────────────────────────────────────────────────────────
class TestQualityCompounderROE:
    """Test 4"""

    def test_all_results_have_roe_gt_15(self, engine):
        """Every company in Quality Compounder must have ROE > 15%."""
        result = engine.preset_quality_compounder()
        roe = pd.to_numeric(result["return_on_equity_pct"], errors="coerce")
        assert (roe > 15.0).all(), (
            f"Some ROE values <= 15%: {result.loc[roe <= 15.0, ['company_id', 'return_on_equity_pct']]}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Test 5 — Quality Compounder: non-Financials all have D/E < 1.0
# ─────────────────────────────────────────────────────────────────────────────
class TestQualityCompounderDE:
    """Test 5"""

    def test_non_financials_have_de_lt_1(self, engine):
        """Non-Financials companies in Quality Compounder must have D/E < 1.0."""
        result = engine.preset_quality_compounder()
        non_fin = result[result["broad_sector"] != "Financials"]
        if len(non_fin) == 0:
            pytest.skip("No non-Financials companies in result")
        de = pd.to_numeric(non_fin["debt_to_equity"], errors="coerce")
        assert (de < 1.0).all(), (
            f"Non-Financials with D/E >= 1.0: "
            f"{non_fin.loc[de >= 1.0, ['company_id', 'debt_to_equity']]}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Test 6 — Quality Compounder: count in 5–50 range
# ─────────────────────────────────────────────────────────────────────────────
class TestQualityCompounderCount:
    """Test 6"""

    def test_result_count_5_to_50(self, engine):
        """Quality Compounder must return 5–50 companies."""
        result = engine.preset_quality_compounder()
        n = len(result)
        assert 5 <= n <= 50, f"Quality Compounder returned {n} companies (expected 5–50)"


# ─────────────────────────────────────────────────────────────────────────────
# Test 7 — All 6 presets return 5–50 companies
# ─────────────────────────────────────────────────────────────────────────────
class TestAllPresetsCount:
    """Test 7"""

    def test_all_presets_return_5_to_50(self, all_presets):
        """Every preset must return between 5 and 50 companies."""
        failures = []
        for name, df in all_presets.items():
            n = len(df)
            if not (5 <= n <= 50):
                failures.append(f"'{name}': {n} companies")
        assert not failures, f"Presets outside 5–50 range: {failures}"


# ─────────────────────────────────────────────────────────────────────────────
# Test 8 — Composite score: all values in [0, 100]
# ─────────────────────────────────────────────────────────────────────────────
class TestCompositeScoreRange:
    """Test 8"""

    def test_composite_score_in_0_100(self, engine):
        """All composite_quality_score values must be in [0, 100]."""
        scores = pd.to_numeric(engine.data["composite_quality_score"], errors="coerce")
        scores = scores.dropna()
        assert len(scores) > 0, "No composite score values found"
        assert (scores >= 0).all() and (scores <= 100).all(), (
            f"Score out of range — min={scores.min():.2f}, max={scores.max():.2f}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Test 9 — Sector-relative score computed for each row
# ─────────────────────────────────────────────────────────────────────────────
class TestSectorRelativeScore:
    """Test 9"""

    def test_sector_relative_score_present(self, engine):
        """sector_relative_score must be non-null for all rows."""
        sr = pd.to_numeric(engine.data["sector_relative_score"], errors="coerce")
        null_count = sr.isna().sum()
        assert null_count == 0, f"{null_count} rows have null sector_relative_score"


# ─────────────────────────────────────────────────────────────────────────────
# Test 10 — Peer percentile: IT Services highest ROE = highest ROE rank
# ─────────────────────────────────────────────────────────────────────────────
class TestITServicesROEPercentile:
    """Test 10"""

    def test_it_services_highest_roe_has_rank_1(self, db_path):
        """The IT Services company with the highest ROE must have percentile_rank = 1.0."""
        try:
            with sqlite3.connect(db_path) as conn:
                rows = conn.execute(
                    """
                    SELECT company_id, value, percentile_rank
                    FROM peer_percentiles
                    WHERE peer_group_name = 'IT Services' AND metric = 'ROE'
                    ORDER BY value DESC
                    """
                ).fetchall()
        except sqlite3.OperationalError:
            pytest.skip("peer_percentiles table does not exist")

        assert len(rows) > 0, "No IT Services ROE rows in peer_percentiles"
        top_company_rank = rows[0][2]   # highest ROE company's percentile rank
        assert top_company_rank == pytest.approx(1.0, abs=0.001), (
            f"IT Services top ROE company has rank {top_company_rank} (expected 1.0)"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Test 11 — FMCG peer group has correct number of companies
# ─────────────────────────────────────────────────────────────────────────────
class TestFMCGPeerGroup:
    """Test 11"""

    def test_fmcg_group_company_count(self, db_path):
        """FMCG peer group must have exactly 7 companies (matching peer_groups table)."""
        with sqlite3.connect(db_path) as conn:
            n_pg = conn.execute(
                "SELECT COUNT(*) FROM peer_groups WHERE peer_group_name = 'FMCG'"
            ).fetchone()[0]
        assert n_pg == 7, f"Expected 7 FMCG companies, found {n_pg}"

        try:
            with sqlite3.connect(db_path) as conn:
                n_pct = conn.execute(
                    """
                    SELECT COUNT(DISTINCT company_id) FROM peer_percentiles
                    WHERE peer_group_name = 'FMCG'
                    """
                ).fetchone()[0]
            assert n_pct == 7, f"peer_percentiles has {n_pct} FMCG companies (expected 7)"
        except sqlite3.OperationalError:
            pytest.skip("peer_percentiles table not populated yet")


# ─────────────────────────────────────────────────────────────────────────────
# Test 12 — D/E filter: Financials sector companies pass unconditionally
# ─────────────────────────────────────────────────────────────────────────────
class TestFinancialsSectorDEExemption:
    """Test 12"""

    def test_financials_not_excluded_by_de_filter(self, engine):
        """Financials sector companies must NOT be excluded when D/E max filter applies."""
        from src.screener.engine import ScreenerEngine
        # Apply a tight D/E filter that all Financials would fail on their own
        result = engine.apply_filters({"debt_to_equity_max": 0.5})
        fin_in_full = engine.data[engine.data["broad_sector"] == "Financials"]
        fin_in_result = result[result["broad_sector"] == "Financials"]
        # All Financials must appear in the filtered result
        assert len(fin_in_result) == len(fin_in_full), (
            f"Expected all {len(fin_in_full)} Financials, got {len(fin_in_result)}"
        )


# ─────────────────────────────────────────────────────────────────────────────
# Test 13 — ICR "Debt Free" passes ICR min filter
# ─────────────────────────────────────────────────────────────────────────────
class TestDebtFreeICRFilter:
    """Test 13"""

    def test_debt_free_companies_pass_icr_filter(self, engine, db_path):
        """
        'Debt Free' ICR label must always pass an ICR min filter.
        Verified using a synthetic DataFrame row injected into apply_filters logic.
        Also confirms the DB contains 'Debt Free' rows across all years.
        """
        # Confirm DB has Debt Free rows (any year)
        with sqlite3.connect(db_path) as conn:
            n_df = conn.execute(
                "SELECT COUNT(*) FROM financial_ratios WHERE icr_label = 'Debt Free'"
            ).fetchone()[0]
        assert n_df > 0, "No 'Debt Free' rows found in financial_ratios"

        # Directly test the filter logic with synthetic engine data
        import pandas as pd
        import copy

        # Build a small test DataFrame with one Debt Free company
        test_df = engine.data.head(3).copy()
        # Force one row to be Debt Free with NULL interest_coverage
        test_df = test_df.reset_index(drop=True)
        test_df.loc[0, "icr_label"]         = "Debt Free"
        test_df.loc[0, "interest_coverage"]  = None

        # Temporarily replace engine._df
        original_df = engine._df
        engine._df  = test_df
        try:
            result = engine.apply_filters({"interest_coverage_min": 100.0})
            debt_free_in_result = result[result["icr_label"] == "Debt Free"]
            assert len(debt_free_in_result) == 1, (
                "Debt Free company was excluded by ICR filter (should always pass)"
            )
        finally:
            engine._df = original_df


# ─────────────────────────────────────────────────────────────────────────────
# Test 14 — peer_comparison.xlsx has exactly 11 sheets
# ─────────────────────────────────────────────────────────────────────────────
class TestPeerComparisonExcel:
    """Test 14"""

    def test_peer_comparison_has_11_sheets(self):
        """output/peer_comparison.xlsx must have exactly 11 sheets."""
        import sys
        sys.path.insert(0, str(Path(__file__).resolve().parents[2]))
        import config
        import openpyxl

        output_path = config.OUTPUT_DIR / "peer_comparison.xlsx"
        if not output_path.exists():
            pytest.skip("peer_comparison.xlsx not generated yet")

        wb = openpyxl.load_workbook(output_path)
        n_sheets = len(wb.sheetnames)
        assert n_sheets == 11, (
            f"peer_comparison.xlsx has {n_sheets} sheets (expected 11): {wb.sheetnames}"
        )
